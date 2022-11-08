import openpyxl
import docx
import os


def main():
    ip = input('输入模板文件名：')
    ip = clean_filename(ip)

    dir_path = input('输入附报文档所在目录，留空表示当前目录：')
    dir_path = clean_filename(dir_path)
    if not dir_path:
        dir_path = os.getcwd()

    output_dir_path = input('输入输出目录，留空表示当前目录：')
    output_dir_path = clean_filename(output_dir_path)
    if not output_dir_path:
        output_dir_path = os.getcwd()

    print('加载工作簿...')
    wb = openpyxl.load_workbook(ip, data_only=True)

    if 'MB' not in wb:
        print('工作簿中不存在MB工作表，请检查模板文件')
        return

    jobs = {}
    sheet = wb['MB']
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        if row[0].value is None or row[1].value is None:
            break

        # job: [
        #     { key: value },
        #     output_filename
        # ]
        jobs[row[0].value] = [{}, row[1].value]
        jobs[row[0].value][1] = row[1].value
        for cell in row[2:]:
            key = sheet.cell(row=1, column=cell.column).value
            jobs[row[0].value][0][key] = cell.value

    print('将替换', len(jobs), '个文档')

    for filename in jobs:
        real_filename = filename
        if not filename.endswith('.docx'):
            real_filename += '.docx'
        print('替换', real_filename)

        if not os.path.exists(os.path.join(dir_path, real_filename)):
            print('文件不存在，跳过')
            continue

        doc = docx.Document(os.path.join(dir_path, real_filename))

        for key in jobs[filename][0]:
            # print('替换', key, '->', jobs[filename][key])
            if jobs[filename][0][key] is None:
                continue
            if jobs[filename][0][key] == "[不替换]":
                continue
            replace_key_in_doc(doc, key, jobs[filename][0][key])

        output_filename = jobs[filename][1]
        if not output_filename.endswith('.docx'):
            output_filename += '.docx'
        doc.save(os.path.join(output_dir_path, output_filename))

def shuttle_text(shuttle):
    t = ''
    for i in shuttle:
        t += i.text
    return t

def replace_key_in_doc(doc, key, value):
    for p in doc.paragraphs:

        begin = 0
        for end in range(len(p.runs)):

            shuttle = p.runs[begin:end+1]

            full_text = shuttle_text(shuttle)
            # print('full_text:', full_text)
            if key in full_text:
                # print('Replace：', key, '->', value)
                # print([i.text for i in shuttle])

                # find the begin
                index = full_text.index(key)
                # print('full_text length', len(full_text), 'index:', index)
                while index >= len(p.runs[begin].text):
                    index -= len(p.runs[begin].text)
                    begin += 1

                shuttle = p.runs[begin:end+1]

                # do replace
                # print('before replace', [i.text for i in shuttle])
                if key in shuttle[0].text:
                    shuttle[0].text = shuttle[0].text.replace(key, value)
                else:
                    replace_begin_index = shuttle_text(shuttle).index(key)
                    replace_end_index = replace_begin_index + len(key)
                    replace_end_index_in_last_run = replace_end_index - len(shuttle_text(shuttle[:-1]))
                    shuttle[0].text = shuttle[0].text[:replace_begin_index] + value

                    # clear middle runs
                    for i in shuttle[1:-1]:
                        i.text = ''

                    # keep last run
                    shuttle[-1].text = shuttle[-1].text[replace_end_index_in_last_run:]

                # print('after replace', [i.text for i in shuttle])

                # set begin to next
                begin = end


def clean_filename(s):
    """
    Return a "cleaned" version of the filename `s`, i.e. with
    leading and trailing whitespace removed, and with internal
    whitespace replaced by underscores.
    """
    s = s.strip()
    if s.startswith('"'):
        s = s[1:]
    if s.endswith('"'):
        s = s[:-1]
    if s.startswith("'"):
        s = s[1:]
    if s.endswith("'"):
        s = s[:-1]
    return s.strip()

if __name__ == '__main__':
    main()
