# coding: utf-8
import openpyxl

# print dict sorted by value
def print_dict_sorted_by_value(d):
    for k, v in sorted(d.items(), key=lambda x: x[1], reverse=False):
        print(v,'\t', k)

def is_empty(s):
    if not s:
        return True
    if s == '' or s == ' ':
        return True
    if s == '--':
        return True
    if s == 'None':
        return True
    if s == '——':
        return True
    return False

# add thousand separator to float numbers
def add_thousand_separator(num):
    try:
        num = float(num)
        return "{:,}".format(num)
    except:
        print('cannot format', num)
        return num

# check if is numbers
def is_number(s):
    try:
        float(s)
        return True
    except:
        return False

def clean_quote_at_begin_and_end(s):
    s = s.strip()
    if s[0] == '"':
        s = s[1:]
    if s[-1] == '"':
        s = s[:-1]
    if s[0] == "'":
        s = s[1:]
    if s[-1] == "'":
        s = s[:-1]
    return s.strip()

print('对比两个Excel文件的差异，将差异写入新的Excel文件中')
print('输入文件名，或脱拽文件到此窗口')
new_wb_name = input('新工作簿：')
new_wb_name = clean_quote_at_begin_and_end(new_wb_name)
old_wb_name = input('旧工作簿：')
old_wb_name = clean_quote_at_begin_and_end(old_wb_name)

fill_rbg_color = '39ceff'
fill_rbg_color_input = input('填充颜色（留空默认39ceff蓝色，输入None不填充颜色）：')
if fill_rbg_color_input.strip():
    fill_rbg_color = fill_rbg_color_input.strip()

print('读取工作簿中...', new_wb_name)
new_wb = openpyxl.load_workbook(new_wb_name, data_only=True)
print('读取工作簿中...', old_wb_name)
old_wb = openpyxl.load_workbook(old_wb_name, data_only=True)

print('对比差异中...')

diff_count = {}
for sheet_name in new_wb.sheetnames:
    if not sheet_name in old_wb.sheetnames:
        continue
    diff_count[sheet_name] = 0
    new_sheet = new_wb[sheet_name]
    old_sheet = old_wb[sheet_name]
    for row in new_sheet:
        for cell in row:
            old_cell = old_sheet.cell(cell.row, cell.column)
            if cell.value != old_cell.value:
                diff_count[sheet_name] += 1
                if is_empty(cell.value) and is_empty(old_cell.value):
                    continue
                print(sheet_name, old_cell.value, '->', cell.value)
                if fill_rbg_color.strip().lower() != 'none':
                    cell.fill = openpyxl.styles.PatternFill("solid", start_color=fill_rbg_color)
                try:
                    old_value = old_cell.value
                    if is_number(old_value):
                        old_value = add_thousand_separator(old_value)
                    cell.comment = openpyxl.comments.Comment(str(old_value), 'python')
                except AttributeError:
                    pass
print('------------------------------------------------------')
print('差异数量：')
print_dict_sorted_by_value(diff_count)
print('------------------------------------------------------')

for sheet_name in old_wb.sheetnames:
    if sheet_name not in new_wb.sheetnames:
        print('旧工作簿中有工作表：', sheet_name, '，但新工作簿中没有，跳过')
        continue
for sheet_name in new_wb.sheetnames:
    if sheet_name not in old_wb.sheetnames:
        print('新工作簿中有工作表：', sheet_name, '，但旧工作簿中没有，跳过')
        continue

new_filename = input('输入新工作表保存文件名：').strip()
if not new_filename.strip():
    new_filename = 'result'
print('保存文件中...')
new_wb.save(clean_quote_at_begin_and_end(new_filename)+'.xlsx')

input('程序结束，按回车键退出')
