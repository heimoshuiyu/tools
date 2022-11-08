import openpyxl
import sys

def main():
    if len(sys.argv) != 2:
        filename = input('使用方法：脱拽文件到此窗口或本文件上：').strip()
        filename = clean_filename(filename)
    else:
        filename = sys.argv[1]

    print('读取文件中...')
    # Open the workbook
    wb = openpyxl.load_workbook(filename)

    total_wb = len(wb.sheetnames)
    for count, sheet in enumerate(wb):
        print('处理工作表中', count+1, '/', total_wb, end='\r')
        for row in sheet:
            for cell in row:
                try:
                    if cell.value is None:
                        continue
                    if not isinstance(cell.value, str):
                        continue
                    if cell.value.startswith("[") and cell.value.endswith("]"):
                        # print(cell.value)
                        text = cell.value[1:-1]
                        sheet_code, cell_code = text.split("$")
                        sheet_name = None
                        for name in wb.sheetnames:
                            if not ' ' in name:
                                continue
                            if sheet_code == name.split(' ')[0]:
                                sheet_name = name
                                break
                        if sheet_name is None:
                            # print("ERROR: Sheet code not found:", sheet_code)
                            continue
                        cell.value = "='%s'!%s" % (sheet_name, cell_code)
                except Exception as e:
                    # print(e)
                    continue
    print()


    # save
    print('保存文件中...')
    wb.save(filename)

def clean_filename(i):
    i = i.strip()
    if i.startswith('"'):
        i = i[1:]
    if i.endswith('"'):
        i = i[:-1]
    if i.startswith("'"):
        i = i[1:]
    if i.endswith("'"):
        i = i[:-1]
    return i.strip()

if __name__ == '__main__':
    main()
